from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from libs.lead_storage import DEFAULT_HEADERS, ExcelLeadWriter


def test_excel_lead_writer_creates_file(tmp_path: Path):
    path = tmp_path / "leads.xlsx"
    writer = ExcelLeadWriter(path)

    lead = {
        "inserted_at": datetime(2024, 9, 24, 15, 45),
        "email": "lead@example.com",
        "first_name": "Lead",
        "last_name": "Example",
        "company": "Example Corp",
        "phone": "+3912345678",
        "subject": "Richiesta preventivo",
        "received_at": datetime(2024, 9, 24, 15, 30),
        "notes": "Body excerpt",
    }

    first_row = writer.append(lead)
    assert first_row == 2

    assert path.exists(), "Workbook should be created"

    second_row = writer.append(lead)
    assert second_row == 3

    # Workbook should contain header + 2 rows
    wb = load_workbook(path)
    data_ws = wb["Leads"]
    assert data_ws.max_row == 3
    assert str(data_ws["A2"].value).startswith("2024-09-24")
    assert data_ws["B2"].value == "lead@example.com"

    summary_ws = wb["Summary"]
    assert summary_ws["A2"].value == "Totale lead"
    assert summary_ws["B2"].value == 2


def test_excel_lead_writer_from_env_custom_headers(monkeypatch, tmp_path: Path):
    target_path = tmp_path / "custom.xlsx"
    monkeypatch.setenv("EXCEL_PATH", str(target_path))
    monkeypatch.setenv("LEADS_XLSX_PATH", str(tmp_path / "legacy.xlsx"))
    monkeypatch.setenv("EXCEL_HEADERS", "email,first_name,inserted_at")

    writer = ExcelLeadWriter.from_env()
    assert writer is not None
    assert writer.path == target_path
    assert list(writer.headers) == ["email", "first_name", "inserted_at"]


def test_excel_lead_writer_from_env_fallback(monkeypatch, tmp_path: Path):
    fallback_path = tmp_path / "fallback.xlsx"
    monkeypatch.delenv("EXCEL_PATH", raising=False)
    monkeypatch.delenv("EXCEL_HEADERS", raising=False)
    monkeypatch.setenv("LEADS_XLSX_PATH", str(fallback_path))

    writer = ExcelLeadWriter.from_env()
    assert writer is not None
    assert writer.path == fallback_path
    assert tuple(writer.headers) == DEFAULT_HEADERS


def test_excel_lead_writer_respects_header_order(tmp_path: Path):
    path = tmp_path / "custom_order.xlsx"
    headers = ["email", "first_name", "inserted_at"]
    writer = ExcelLeadWriter(path, headers=headers)

    lead = {
        "inserted_at": "2024-01-15T10:00:00",
        "email": "lead@example.com",
        "first_name": "Lead",
    }

    writer.append(lead)

    wb = load_workbook(path)
    data_ws = wb["Leads"]

    assert [cell.value for cell in data_ws[1][: len(headers)]] == headers
    assert data_ws["A2"].value == "lead@example.com"
    assert data_ws["B2"].value == "Lead"
    assert str(data_ws["C2"].value).startswith("2024-01-15")
