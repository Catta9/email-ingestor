from __future__ import annotations

import os
from collections import Counter
from collections.abc import Mapping as MappingABC
from collections.abc import Sequence as SequenceABC
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping, Optional, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_HEADERS: tuple[str, ...] = (
    "inserted_at",
    "email",
    "first_name",
    "last_name",
    "company",
    "phone",
    "subject",
    "received_at",
    "notes",
)

DATA_SHEET_TITLE = "Leads"
SUMMARY_SHEET_TITLE = "Summary"
DEFAULT_TABLE_NAME = "LeadsTable"


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _coerce_datetime(value: object) -> object:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str) and value:
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            return value
    return value


def _header_index(headers: Sequence[str], target: str) -> Optional[int]:
    lookup = {name.lower(): idx for idx, name in enumerate(headers, start=1)}
    return lookup.get(target.lower())


def _prepare_headers(headers: Iterable[object]) -> list[str]:
    header_list = [str(header).strip() for header in headers if str(header).strip()]
    return header_list or list(DEFAULT_HEADERS)


def _extract_headers(ws) -> list[str]:
    if ws.max_row < 1:
        return []
    header_values = []
    for cell in ws[1]:
        if cell.value is None:
            continue
        header_values.append(str(cell.value).strip())
    return [value for value in header_values if value]


def _coerce_header_value(header: str, value: object) -> object:
    if header.lower() in {"inserted_at", "received_at"}:
        return _coerce_datetime(value)
    return value


def _row_from_mapping(headers: Sequence[str], row: MappingABC[str, object]) -> list[object]:
    normalized = {str(key).lower(): value for key, value in row.items()}
    return [
        _coerce_header_value(header, normalized.get(header.lower()))
        for header in headers
    ]


def _row_from_sequence(headers: Sequence[str], row: SequenceABC) -> list[object]:
    sequence = list(row)
    values: list[object] = []
    for index, header in enumerate(headers):
        value = sequence[index] if index < len(sequence) else None
        values.append(_coerce_header_value(header, value))
    return values


def _row_from_any(headers: Sequence[str], row: object) -> list[object]:
    if isinstance(row, MappingABC):
        return _row_from_mapping(headers, row)
    if isinstance(row, SequenceABC) and not isinstance(row, (str, bytes, bytearray)):
        return _row_from_sequence(headers, row)
    return [
        _coerce_header_value(header, getattr(row, header, None))
        for header in headers
    ]


def _apply_header_style(ws, header_count: int) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, header_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        column_cells = list(column_cells)
        if not column_cells:
            continue
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = cell.value
            length = len(str(value)) if value is not None else 0
            max_length = max(max_length, length)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def _ensure_table(ws, header_count: int, *, table_name: str) -> None:
    max_row = ws.max_row or 1
    max_col = ws.max_column or header_count
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    if table_name in ws.tables:
        ws.tables[table_name].ref = ref
        return

    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _email_domain(value: object) -> Optional[str]:
    if not value:
        return None
    email = str(value).strip().lower()
    if "@" not in email:
        return None
    return email.rsplit("@", 1)[1]


def _collect_column_values(ws, column_index: int) -> list[object]:
    values = []
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=column_index).value
        if value:
            values.append(value)
    return values


def _style_header_row(ws, row_idx: int, column_count: int) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alignment = Alignment(horizontal="center")
    for col in range(1, column_count + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment


def _update_summary_sheet(
    wb: Workbook,
    headers: Sequence[str],
    *,
    data_sheet_name: str = DATA_SHEET_TITLE,
    summary_sheet_name: str = SUMMARY_SHEET_TITLE,
) -> None:
    header_list = list(headers)
    email_idx = _header_index(header_list, "email")
    company_idx = _header_index(header_list, "company") or _header_index(header_list, "org")

    if summary_sheet_name in wb.sheetnames:
        summary_ws = wb[summary_sheet_name]
        summary_ws.delete_rows(1, summary_ws.max_row)
    else:
        summary_ws = wb.create_sheet(summary_sheet_name, 0)

    data_ws = wb[data_sheet_name]
    total_leads = max(data_ws.max_row - 1, 0)
    summary_ws.append(["Metric", "Value"])
    _style_header_row(summary_ws, 1, 2)
    summary_ws.freeze_panes = "A2"
    summary_ws.column_dimensions["A"].width = 28
    summary_ws.column_dimensions["B"].width = 40

    summary_ws.append(["Totale lead", total_leads])

    if company_idx:
        companies = {
            str(value).strip().lower()
            for value in _collect_column_values(data_ws, company_idx)
            if str(value).strip()
        }
        summary_ws.append(["Aziende uniche", len(companies)])

    if email_idx:
        domain_counter: Counter[str] = Counter()
        for value in _collect_column_values(data_ws, email_idx):
            domain = _email_domain(value)
            if domain:
                domain_counter[domain] += 1

        if domain_counter:
            summary_ws.append([])
            summary_ws.append(["Dominio email", "Lead"])
            header_row = summary_ws.max_row
            _style_header_row(summary_ws, header_row, 2)
            for domain, count in domain_counter.most_common(10):
                summary_ws.append([domain, count])

    summary_ws.append([])
    summary_ws.append(["Ultimo aggiornamento", datetime.utcnow()])

    _auto_fit_columns(summary_ws)

    idx = wb.sheetnames.index(summary_sheet_name)
    if idx != 0:
        wb.move_sheet(summary_ws, -idx)

    wb.active = wb.sheetnames.index(data_sheet_name)


def _finalize_data_sheet(ws, header_count: int, table_name: str) -> None:
    ws.freeze_panes = "A2"
    _apply_header_style(ws, header_count)
    _ensure_table(ws, header_count, table_name=table_name)
    _auto_fit_columns(ws)


def build_structured_workbook(
    headers: Sequence[str],
    rows: Iterable[Sequence[object] | Mapping[str, object]],
    *,
    data_sheet_name: str = DATA_SHEET_TITLE,
    summary_sheet_name: str = SUMMARY_SHEET_TITLE,
    table_name: str = DEFAULT_TABLE_NAME,
) -> Workbook:
    wb = Workbook()
    data_ws = wb.active
    data_ws.title = data_sheet_name
    header_list = _prepare_headers(headers)
    data_ws.append(header_list)
    for row in rows:
        data_ws.append(_row_from_any(header_list, row))

    _finalize_data_sheet(data_ws, len(header_list), table_name)
    _update_summary_sheet(
        wb,
        header_list,
        data_sheet_name=data_sheet_name,
        summary_sheet_name=summary_sheet_name,
    )
    return wb


@dataclass
class ExcelLeadWriter:
    """Append lead rows to an Excel workbook, creating it if needed."""

    path: Path
    headers: Iterable[str] = DEFAULT_HEADERS

    @classmethod
    def from_env(cls) -> "ExcelLeadWriter | None":
        primary = (os.getenv("EXCEL_PATH") or "").strip()
        fallback_env = (os.getenv("LEADS_XLSX_PATH") or "").strip()
        fallback = fallback_env or "data/leads.xlsx"
        target = primary or fallback
        if not target:
            return None
        headers_env = os.getenv("EXCEL_HEADERS")
        if headers_env:
            headers = _prepare_headers(headers_env.split(","))
        else:
            headers = list(DEFAULT_HEADERS)
        return cls(Path(target), headers=tuple(headers))

    def append(self, lead: dict[str, Optional[str | datetime]]) -> int:
        desired_headers = _prepare_headers(self.headers)
        _ensure_parent(self.path)
        if self.path.exists():
            wb = load_workbook(self.path)
            data_ws = wb[DATA_SHEET_TITLE] if DATA_SHEET_TITLE in wb.sheetnames else wb.active
            existing_headers = _extract_headers(data_ws)
            headers = existing_headers or desired_headers
            if not existing_headers:
                first_row = list(data_ws[1]) if data_ws.max_row else []
                if first_row and all(cell.value is None for cell in first_row):
                    data_ws.delete_rows(1, 1)
                data_ws.append(headers)
        else:
            wb = Workbook()
            data_ws = wb.active
            data_ws.title = DATA_SHEET_TITLE
            headers = desired_headers
            data_ws.append(headers)

        row = _row_from_any(headers, lead)

        data_ws.append(row)
        _finalize_data_sheet(data_ws, len(headers), DEFAULT_TABLE_NAME)
        _update_summary_sheet(wb, headers)
        wb.save(self.path)
        return data_ws.max_row
